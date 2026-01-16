"""
Excel exporter for tennis match odds data.
Handles exporting scraped data to Excel format with formatting and multiple sheets.
"""

from typing import List, Dict, Any
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import get_logger, ensure_directory


class ExcelExporter:
    """
    Export tennis match data to Excel format with professional formatting.
    Supports multiple sheets, auto-fitting, and styled output.
    """
    
    def __init__(self, output_dir: str = 'data'):
        """
        Initialize Excel exporter.
        
        Args:
            output_dir: Directory to save Excel files
        """
        self.output_dir = ensure_directory(output_dir)
        self.logger = get_logger('exporter.excel')
    
    def export(
        self,
        matches: List[Dict[str, Any]],
        filename: str,
        include_summary: bool = True,
        include_calculations: bool = True
    ) -> str:
        """
        Export matches to Excel file with formatting.
        
        Args:
            matches: List of match dictionaries
            filename: Output filename (without path)
            include_summary: Add summary sheet
            include_calculations: Add calculated fields
            
        Returns:
            Full path to created Excel file
        """
        if not matches:
            self.logger.warning("No matches to export")
            return ""
        
        try:
            # Build output path
            filepath = self.output_dir / filename
            if not filepath.suffix:
                filepath = filepath.with_suffix('.xlsx')
            
            # Convert to DataFrame
            df = pd.DataFrame(matches)
            
            # Add calculated fields
            if include_calculations:
                df = self._add_calculations(df)
            
            # Order columns
            df = self._order_columns(df)
            
            # Create Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write main data sheet
                df.to_excel(writer, sheet_name='Match Data', index=False)
                
                # Write summary sheet if requested
                if include_summary:
                    summary_df = self._create_summary(df)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Apply formatting
            self._apply_formatting(filepath)
            
            self.logger.info(f"Exported {len(matches)} matches to {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _add_calculations(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Add calculated fields to DataFrame.
        
        Args:
            df: Input DataFrame
            
        Returns:
            DataFrame with additional columns
        """
        try:
            # Implied probability for each player
            if 'odds_player1' in df.columns:
                df['implied_prob1'] = df['odds_player1'].apply(
                    lambda x: round((1 / x) * 100, 2) if x and x > 0 else None
                )
            
            if 'odds_player2' in df.columns:
                df['implied_prob2'] = df['odds_player2'].apply(
                    lambda x: round((1 / x) * 100, 2) if x and x > 0 else None
                )
            
            # Bookmaker margin
            if 'odds_player1' in df.columns and 'odds_player2' in df.columns:
                df['bookmaker_margin'] = df.apply(
                    lambda row: self._calculate_margin(
                        row['odds_player1'],
                        row['odds_player2']
                    ),
                    axis=1
                )
        
        except Exception as e:
            self.logger.warning(f"Error adding calculations: {e}")
        
        return df
    
    @staticmethod
    def _calculate_margin(odds1: float, odds2: float) -> float:
        """Calculate bookmaker margin."""
        if not odds1 or not odds2 or odds1 <= 0 or odds2 <= 0:
            return None
        
        prob1 = 1 / odds1
        prob2 = 1 / odds2
        margin = ((prob1 + prob2 - 1) * 100)
        
        return round(margin, 2)
    
    def _order_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Order DataFrame columns for better readability.
        
        Args:
            df: Input DataFrame
            
        Returns:
            DataFrame with reordered columns
        """
        preferred_order = [
            'timestamp', 'bookmaker', 'tournament',
            'player1', 'player2',
            'odds_player1', 'odds_player2',
            'implied_prob1', 'implied_prob2',
            'bookmaker_margin',
            'match_date', 'match_time', 'url'
        ]
        
        # Keep only columns that exist
        available_cols = [col for col in preferred_order if col in df.columns]
        
        # Add any remaining columns
        remaining_cols = [col for col in df.columns if col not in available_cols]
        
        return df[available_cols + remaining_cols]
    
    def _create_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Create summary statistics DataFrame.
        
        Args:
            df: Input DataFrame
            
        Returns:
            Summary DataFrame
        """
        summary_data = []
        
        # Basic counts
        summary_data.append({
            'Metric': 'Total Matches',
            'Value': len(df)
        })
        
        if 'tournament' in df.columns:
            summary_data.append({
                'Metric': 'Unique Tournaments',
                'Value': df['tournament'].nunique()
            })
        
        if 'bookmaker' in df.columns:
            summary_data.append({
                'Metric': 'Bookmakers',
                'Value': df['bookmaker'].nunique()
            })
        
        # Odds statistics
        if 'odds_player1' in df.columns and 'odds_player2' in df.columns:
            all_odds = pd.concat([df['odds_player1'], df['odds_player2']])
            
            summary_data.extend([
                {'Metric': 'Average Odds (All)', 'Value': round(all_odds.mean(), 2)},
                {'Metric': 'Min Odds', 'Value': round(all_odds.min(), 2)},
                {'Metric': 'Max Odds', 'Value': round(all_odds.max(), 2)}
            ])
        
        # Margin statistics
        if 'bookmaker_margin' in df.columns:
            summary_data.extend([
                {'Metric': 'Average Margin (%)', 'Value': round(df['bookmaker_margin'].mean(), 2)},
                {'Metric': 'Min Margin (%)', 'Value': round(df['bookmaker_margin'].min(), 2)},
                {'Metric': 'Max Margin (%)', 'Value': round(df['bookmaker_margin'].max(), 2)}
            ])
        
        return pd.DataFrame(summary_data)
    
    def _apply_formatting(self, filepath: Path) -> None:
        """
        Apply Excel formatting to workbook.
        
        Args:
            filepath: Path to Excel file
        """
        try:
            wb = load_workbook(filepath)
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self._format_sheet(ws)
            
            wb.save(filepath)
            
        except Exception as e:
            self.logger.warning(f"Error applying formatting: {e}")
    
    def _format_sheet(self, ws) -> None:
        """
        Format a single worksheet.
        
        Args:
            ws: Worksheet object
        """
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Apply borders to all cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
    
    def export_by_tournament(
        self,
        matches: List[Dict[str, Any]],
        filename: str
    ) -> str:
        """
        Export matches grouped by tournament (one sheet per tournament).
        
        Args:
            matches: List of match dictionaries
            filename: Output filename
            
        Returns:
            Full path to created Excel file
        """
        if not matches:
            return ""
        
        try:
            filepath = self.output_dir / filename
            if not filepath.suffix:
                filepath = filepath.with_suffix('.xlsx')
            
            df = pd.DataFrame(matches)
            
            # Group by tournament
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                if 'tournament' in df.columns:
                    for tournament, group in df.groupby('tournament'):
                        # Sanitize sheet name (max 31 chars, no special chars)
                        sheet_name = str(tournament)[:31]
                        sheet_name = sheet_name.replace('/', '-').replace('\\', '-')
                        
                        group.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    df.to_excel(writer, sheet_name='All Matches', index=False)
            
            # Apply formatting
            self._apply_formatting(filepath)
            
            self.logger.info(f"Exported to {filepath} with tournament grouping")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Error exporting by tournament: {e}")
            raise


# Example usage
if __name__ == '__main__':
    # Sample data
    sample_matches = [
        {
            'timestamp': '2025-01-15 10:00:00',
            'tournament': 'ATP Australian Open',
            'player1': 'Djokovic N.',
            'player2': 'Federer R.',
            'odds_player1': 1.50,
            'odds_player2': 2.80,
            'bookmaker': 'oddsportal',
            'match_date': '2025-01-20',
            'match_time': '14:00'
        },
        {
            'timestamp': '2025-01-15 10:01:00',
            'tournament': 'ATP Australian Open',
            'player1': 'Nadal R.',
            'player2': 'Murray A.',
            'odds_player1': 1.75,
            'odds_player2': 2.20,
            'bookmaker': 'oddsportal',
            'match_date': '2025-01-20',
            'match_time': '16:00'
        }
    ]
    
    exporter = ExcelExporter()
    output_file = exporter.export(sample_matches, 'test_export.xlsx')
    print(f"Exported to: {output_file}")
